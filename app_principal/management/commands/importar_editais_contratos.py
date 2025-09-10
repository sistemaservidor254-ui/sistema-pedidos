from django.core.management.base import BaseCommand, CommandError
from app_principal.models import Edital, Fornecedor, ARPContrato
from openpyxl import load_workbook
from decimal import Decimal
import re

def parse_brl(n):
    if n is None or n == '':
        return None
    s = str(n).strip()
    s = s.replace('.', '').replace(',', '.')
    return Decimal(s)

def clean_cnpj(cnpj):
    if not cnpj: return ''
    return re.sub(r'\D+', '', str(cnpj))

class Command(BaseCommand):
    help = "Importa Editais e ARPs a partir de uma planilha .xlsx"

    def add_arguments(self, parser):
        parser.add_argument('arquivo', type=str, help='Caminho do .xlsx com abas: editais, contratos')

    def handle(self, *args, **kwargs):
        wb = load_workbook(kwargs['arquivo'])
        # Aba 'editais': numero | processo | pregao | tipo(ENFERMAGEM/MEDICAMENTOS) | objeto
        if 'editais' in wb.sheetnames:
            sh = wb['editais']
            for row in sh.iter_rows(min_row=2, values_only=True):
                numero, processo, pregao, tipo, objeto = row[:5]
                if not numero: continue
                edital, _ = Edital.objects.get_or_create(numero=str(numero).strip(), defaults={
                    'processo': (processo or '').strip(),
                    'pregao': (pregao or '').strip(),
                    'tipo': str(tipo).strip().upper(),
                    'objeto': objeto or ''
                })
                # Atualiza se já existir
                edital.processo = (processo or edital.processo or '').strip()
                edital.pregao = (pregao or edital.pregao or '').strip()
                edital.tipo = str(tipo).strip().upper() or edital.tipo
                edital.objeto = objeto or edital.objeto
                edital.save()

        # Aba 'contratos': edital_numero | arp_numero | fornecedor_razao | cnpj | valor_total | observacoes
        if 'contratos' in wb.sheetnames:
            sh = wb['contratos']
            for row in sh.iter_rows(min_row=2, values_only=True):
                ed_num, arp_num, razao, cnpj, valor, obs = row[:6]
                if not ed_num or not arp_num or not razao: continue
                try:
                    edital = Edital.objects.get(numero=str(ed_num).strip())
                except Edital.DoesNotExist:
                    raise CommandError(f"Edital {ed_num} não encontrado.")
                cnpj_num = clean_cnpj(cnpj)
                forn, _ = Fornecedor.objects.get_or_create(
                    cnpj=cnpj_num or f"SEM-CNPJ-{razao}",
                    defaults={'razao_social': str(razao).strip()}
                )
                forn.razao_social = str(razao).strip()
                forn.save()
                arp, _ = ARPContrato.objects.get_or_create(
                    numero_arp=str(arp_num).strip(),
                    edital=edital,
                    fornecedor=forn,
                    defaults={'valor_total': parse_brl(valor) if valor else None, 'observacoes': obs or ''}
                )
                if valor:
                    arp.valor_total = parse_brl(valor)
                if obs:
                    arp.observacoes = obs
                arp.save()

        self.stdout.write(self.style.SUCCESS('Editais e ARPs importados/atualizados com sucesso.'))