from django.core.management.base import BaseCommand
from openpyxl import Workbook

class Command(BaseCommand):
    help = "Gera modelos de planilhas .xlsx para importação"

    def handle(self, *args, **kwargs):
        # editais + contratos
        wb = Workbook()
        sh1 = wb.active
        sh1.title = 'editais'
        sh1.append(['numero', 'processo', 'pregao', 'tipo', 'objeto'])
        sh2 = wb.create_sheet('contratos')
        sh2.append(['edital_numero', 'arp_numero', 'fornecedor_razao', 'cnpj', 'valor_total', 'observacoes'])
        wb.save('modelo_editais_contratos.xlsx')

        # itens
        wb2 = Workbook()
        sh = wb2.active
        sh.title = 'itens'
        sh.append([
            'edital_numero','codigo','descricao','unidade','quantidade','valor_unitario','valor_total',
            'fornecedor_cnpj','fornecedor_razao','arp_numero','centro_custo','proponente_codigo','proponente_nome'
        ])
        wb2.save('modelo_itens.xlsx')

        self.stdout.write(self.style.SUCCESS('Modelos gerados: modelo_editais_contratos.xlsx, modelo_itens.xlsx'))