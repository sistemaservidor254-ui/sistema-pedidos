from django.core.management.base import BaseCommand, CommandError
from app_principal.models import Edital, Fornecedor, ARPContrato, ItemLicitado
from openpyxl import load_workbook
from decimal import Decimal
import re

def parse_brl(n):
    if n is None or n == '':
        return None
    s = str(n).strip()
    s = s.replace('.', '').replace(',', '.')
    return Decimal(s)

def clean_cnpj(cnpj):
    if not cnpj: return ''
    return re.sub(r'\D+', '', str(cnpj))

class Command(BaseCommand):
    help = "Importa itens licitados de uma planilha .xlsx"

    def add_arguments(self, parser):
        parser.add_argument('arquivo', type=str, help='Caminho do .xlsx com aba itens')

    def handle(self, *args, **kwargs):
        wb = load_workbook(kwargs['arquivo'])
        if 'itens' not in wb.sheetnames:
            raise CommandError("A planilha precisa ter uma aba chamada 'itens'.")

        sh = wb['itens']
        # Colunas esperadas (linha 1 como cabeçalho):
        # edital_numero | codigo | descricao | unidade | quantidade | valor_unitario | valor_total | fornecedor_cnpj | fornecedor_razao | arp_numero | centro_custo | proponente_codigo | proponente_nome
        header = [c.value for c in sh[1]]
        idx = {name: header.index(name) for name in header if name}

        required = ['edital_numero', 'descricao', 'quantidade']
        for r in required:
            if r not in idx:
                raise CommandError(f"Coluna obrigatória ausente: {r}")

        count = 0
        for row in sh.iter_rows(min_row=2, values_only=True):
            if not row[idx['edital_numero']]:
                continue
            ed_num = str(row[idx['edital_numero']]).strip()
            edital = Edital.objects.get(numero=ed_num)

            codigo = str(row[idx['codigo']]).strip() if 'codigo' in idx and row[idx['codigo']] else ''
            descricao = str(row[idx['descricao']]).strip()
            unidade = str(row[idx['unidade']]).strip() if 'unidade' in idx and row[idx['unidade']] else ''
            quantidade = int(row[idx['quantidade']] or 0)

            vu = parse_brl(row[idx['valor_unitario']]) if 'valor_unitario' in idx else None
            vt = parse_brl(row[idx['valor_total']]) if 'valor_total' in idx else None

            forn = None
            contrato = None
            if 'fornecedor_cnpj' in idx or 'fornecedor_razao' in idx:
                cnpj = clean_cnpj(row[idx.get('fornecedor_cnpj')]) if 'fornecedor_cnpj' in idx else ''
                razao = str(row[idx.get('fornecedor_razao')] or '').strip()
                if cnpj or razao:
                    forn, _ = Fornecedor.objects.get_or_create(
                        cnpj=cnpj or f"SEM-CNPJ-{razao or 'DESCONHECIDO'}",
                        defaults={'razao_social': razao or 'DESCONHECIDO'}
                    )
                    if razao:
                        forn.razao_social = razao
                        forn.save()

            if 'arp_numero' in idx and row[idx['arp_numero']]:
                arp_num = str(row[idx['arp_numero']]).strip()
                if forn:
                    contrato = ARPContrato.objects.filter(numero_arp=arp_num, edital=edital, fornecedor=forn).first()
                else:
                    contrato = ARPContrato.objects.filter(numero_arp=arp_num, edital=edital).first()

            item = ItemLicitado.objects.create(
                edital=edital,
                fornecedor=forn,
                contrato=contrato,
                codigo=codigo,
                descricao=descricao,
                unidade=unidade,
                quantidade=quantidade,
                valor_unitario=vu,
                valor_total=vt,
                centro_custo=str(row[idx['centro_custo']]).strip() if 'centro_custo' in idx and row[idx['centro_custo']] else '',
                proponente_codigo=str(row[idx['proponente_codigo']]).strip() if 'proponente_codigo' in idx and row[idx['proponente_codigo']] else '',
                proponente_nome=str(row[idx['proponente_nome']]).strip() if 'proponente_nome' in idx and row[idx['proponente_nome']] else '',
            )
            count += 1

        self.stdout.write(self.style.SUCCESS(f'{count} itens importados com sucesso.'))